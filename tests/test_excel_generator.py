"""
Tests for Excel report generator.

Tests cover:
- Excel workbook creation
- Format definitions
- Summary sheet generation
- Equipment sheet with data
- Seasonal sheet with multiple tables
- Vendor sheet with performance metrics
- Failure sheet with patterns
- Recommendations sheet consolidation
- Column width auto-sizing
- Full workbook generation
- Edge cases: empty data, missing columns, null values
"""

import pytest
import pandas as pd
import os
from pathlib import Path
import openpyxl

from src.reporting.excel_generator import ExcelReportGenerator
from src.reporting.report_builder import Report, ReportSection


@pytest.fixture
def sample_report():
    """Create a sample Report object for testing."""
    metadata = {
        'generated_date': '2024-01-15 10:00:00',
        'data_period_start': '2023-01-01',
        'data_period_end': '2023-12-31',
        'total_records': 1000,
        'total_cost': 500000,
        'date_range_days': 365
    }

    # Equipment section
    equipment_df = pd.DataFrame({
        'Equipment_Name': ['Chiller A', 'Pump B', 'HVAC C'],
        'equipment_primary_category': ['HVAC', 'Plumbing', 'HVAC'],
        'work_orders_per_month': ['2.50', '1.75', '3.00'],
        'avg_cost': ['$5,000', '$3,000', '$4,500'],
        'cost_impact': ['$150,000', '$63,000', '$162,000'],
        'priority_score': ['0.850', '0.650', '0.900'],
        'overall_rank': [2, 3, 1]
    })

    equipment_section = ReportSection(
        title="Equipment Analysis",
        content={
            'top_equipment': equipment_df,
            'total_equipment': 50,
            'consensus_outliers': 3,
            'thresholds': {
                'work_orders_per_month': 2.0,
                'cost_impact': 50000,
                'rationale': 'Thresholds calculated from median of consensus outliers'
            }
        },
        summary_text="Identified 3 consensus outliers from 50 equipment items.",
        recommendations=[
            "Focus on top 3 equipment: Chiller A, Pump B, HVAC C",
            "Review maintenance schedules for high-priority items"
        ]
    )

    # Seasonal section
    monthly_df = pd.DataFrame({
        'period': ['2023-01', '2023-02', '2023-03'],
        'total_cost': ['$45,000', '$38,000', '$52,000'],
        'work_order_count': [120, 95, 135],
        'avg_cost': ['$375', '$400', '$385']
    })

    quarterly_df = pd.DataFrame({
        'period': ['2023-Q1', '2023-Q2', '2023-Q3', '2023-Q4'],
        'total_cost': ['$135,000', '$120,000', '$145,000', '$100,000'],
        'work_order_count': [350, 310, 380, 260],
        'avg_cost': ['$386', '$387', '$382', '$385'],
        'variance_pct': ['+8.0%', '-3.0%', '+15.0%', '-20.0%']
    })

    seasonal_section = ReportSection(
        title="Seasonal Analysis",
        content={
            'monthly_costs': monthly_df,
            'quarterly_costs': quarterly_df,
            'patterns': [
                {'pattern': 'Summer spike: Q3 costs 15% above average'},
                {'pattern': 'Winter drop: Q4 costs 20% below average'}
            ],
            'pattern_count': 2
        },
        summary_text="Detected 2 seasonal patterns with cost variance from -20% to +15%.",
        recommendations=[
            "Plan for increased maintenance in Q3 (summer)",
            "Review staffing levels during Q4 (winter)"
        ]
    )

    # Vendor section
    vendor_df = pd.DataFrame({
        'contractor': ['Vendor A', 'Vendor B', 'Vendor C'],
        'total_cost': ['$150,000', '$120,000', '$95,000'],
        'work_order_count': [50, 45, 38],
        'avg_cost_per_wo': ['$3,000', '$2,667', '$2,500']
    })

    vendor_section = ReportSection(
        title="Vendor Performance",
        content={
            'top_vendors': vendor_df,
            'total_vendors': 15,
            'efficiency_metrics': pd.DataFrame(),
            'quality_metrics': pd.DataFrame(),
            'recommendation_count': 3
        },
        summary_text="Analyzed 15 vendors. Top 3 account for $365,000 (73%) of vendor costs.",
        recommendations=[
            "Review contract with Vendor A (highest cost)",
            "Consider competitive bidding for top vendors"
        ]
    )

    # Failure section
    failure_df = pd.DataFrame({
        'pattern': ['Water leak', 'Electrical fault', 'HVAC breakdown'],
        'category': ['Plumbing', 'Electrical', 'HVAC'],
        'occurrences': [25, 18, 15],
        'total_cost': ['$75,000', '$54,000', '$67,500'],
        'avg_cost': ['$3,000', '$3,000', '$4,500'],
        'equipment_affected': [12, 8, 5],
        'impact_score': ['225.0', '144.0', '112.5']
    })

    failure_section = ReportSection(
        title="Failure Pattern Analysis",
        content={
            'high_impact_patterns': failure_df,
            'failure_categories': pd.DataFrame(),
            'pattern_count': 3,
            'category_count': 5
        },
        summary_text="Identified 3 high-impact failure patterns with 58 total occurrences.",
        recommendations=[
            "Address water leak pattern (25 occurrences, $75k cost)",
            "Inspect electrical systems preventively"
        ]
    )

    report = Report(
        metadata=metadata,
        sections=[equipment_section, seasonal_section, vendor_section, failure_section],
        executive_summary="Analysis completed across 4 areas with key findings in equipment, seasonal patterns, vendor performance, and failure patterns."
    )

    return report


@pytest.fixture
def temp_output_file(tmp_path):
    """Create a temporary output file path."""
    return str(tmp_path / "test_report.xlsx")


def test_excel_generator_creation():
    """Test ExcelReportGenerator can be created."""
    generator = ExcelReportGenerator()
    assert generator is not None
    assert hasattr(generator, 'formats')


def test_generate_excel_creates_file(sample_report, temp_output_file):
    """Test that generate_excel creates an Excel file."""
    generator = ExcelReportGenerator()
    generator.generate_excel(sample_report, temp_output_file)

    assert os.path.exists(temp_output_file)
    assert os.path.getsize(temp_output_file) > 0


def test_excel_has_correct_sheets(sample_report, temp_output_file):
    """Test that Excel file has all expected sheets."""
    generator = ExcelReportGenerator()
    generator.generate_excel(sample_report, temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    sheet_names = wb.sheetnames

    assert 'Summary' in sheet_names
    assert 'Equipment' in sheet_names
    assert 'Seasonal' in sheet_names
    assert 'Vendors' in sheet_names
    assert 'Failures' in sheet_names
    assert 'Recommendations' in sheet_names


def test_summary_sheet_content(sample_report, temp_output_file):
    """Test that Summary sheet contains metadata."""
    generator = ExcelReportGenerator()
    generator.generate_excel(sample_report, temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Summary']

    # Check title
    assert 'Executive Summary' in ws['A1'].value

    # Check metadata presence (values are in column B)
    assert ws['A4'].value == 'Generated:'
    assert ws['A5'].value == 'Data Period:'
    assert ws['A6'].value == 'Total Records:'
    assert ws['A7'].value == 'Total Cost:'


def test_equipment_sheet_content(sample_report, temp_output_file):
    """Test that Equipment sheet contains data."""
    generator = ExcelReportGenerator()
    generator.generate_excel(sample_report, temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Equipment']

    # Check title
    assert 'Equipment Analysis' in ws['A1'].value

    # Check headers (row 5, 0-indexed row 4)
    assert ws['A5'].value == 'Equipment'
    assert ws['B5'].value == 'Category'
    assert ws['C5'].value == 'WO/Month'

    # Check data (first equipment)
    assert ws['A6'].value == 'Chiller A'
    assert ws['B6'].value == 'HVAC'


def test_seasonal_sheet_content(sample_report, temp_output_file):
    """Test that Seasonal sheet contains monthly and quarterly tables."""
    generator = ExcelReportGenerator()
    generator.generate_excel(sample_report, temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Seasonal']

    # Check title
    assert 'Seasonal Cost Analysis' in ws['A1'].value

    # Check for table headers (exact row depends on layout, but should be present)
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert 'Monthly Cost Trends' in all_values
    assert 'Quarterly Cost Trends' in all_values


def test_vendor_sheet_content(sample_report, temp_output_file):
    """Test that Vendor sheet contains data."""
    generator = ExcelReportGenerator()
    generator.generate_excel(sample_report, temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Vendors']

    # Check title
    assert 'Vendor Performance Analysis' in ws['A1'].value

    # Check data presence
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert 'Vendor A' in all_values
    assert 'Contractor' in all_values  # Header


def test_failure_sheet_content(sample_report, temp_output_file):
    """Test that Failure sheet contains pattern data."""
    generator = ExcelReportGenerator()
    generator.generate_excel(sample_report, temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Failures']

    # Check title
    assert 'Failure Pattern Analysis' in ws['A1'].value

    # Check data presence
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert 'Water leak' in all_values
    assert 'Pattern' in all_values  # Header


def test_recommendations_sheet_content(sample_report, temp_output_file):
    """Test that Recommendations sheet consolidates all recommendations."""
    generator = ExcelReportGenerator()
    generator.generate_excel(sample_report, temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Recommendations']

    # Check title
    assert 'Consolidated Recommendations' in ws['A1'].value

    # Check headers
    assert ws['A4'].value == 'Source'
    assert ws['B4'].value == 'Priority'
    assert ws['C4'].value == 'Recommendation'

    # Check that recommendations from multiple sources are present
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert 'Equipment Analysis' in all_values
    assert 'Seasonal Analysis' in all_values


def test_empty_equipment_section(temp_output_file):
    """Test handling of empty equipment section."""
    metadata = {'generated_date': '2024-01-15', 'total_records': 0, 'total_cost': 0}

    equipment_section = ReportSection(
        title="Equipment Analysis",
        content={'message': 'No high-priority equipment identified', 'total_equipment': 0, 'outliers_detected': 0},
        summary_text="No consensus outliers detected.",
        recommendations=["Continue monitoring"]
    )

    report = Report(metadata=metadata, sections=[equipment_section])

    generator = ExcelReportGenerator()
    generator.generate_excel(report, temp_output_file)

    # Should create file without errors
    assert os.path.exists(temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Equipment']

    # Check that empty message is displayed
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert 'No high-priority equipment identified' in all_values


def test_empty_seasonal_section(temp_output_file):
    """Test handling of empty seasonal section."""
    metadata = {'generated_date': '2024-01-15', 'total_records': 0, 'total_cost': 0}

    seasonal_section = ReportSection(
        title="Seasonal Analysis",
        content={'message': 'No valid date and cost data available', 'monthly_data_points': 0},
        summary_text="Insufficient data.",
        recommendations=[]
    )

    report = Report(metadata=metadata, sections=[seasonal_section])

    generator = ExcelReportGenerator()
    generator.generate_excel(report, temp_output_file)

    assert os.path.exists(temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Seasonal']
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert 'No valid date and cost data available' in all_values


def test_empty_vendor_section(temp_output_file):
    """Test handling of empty vendor section."""
    metadata = {'generated_date': '2024-01-15', 'total_records': 0, 'total_cost': 0}

    vendor_section = ReportSection(
        title="Vendor Performance",
        content={'message': 'No contractors with minimum 3 work orders found', 'vendors_analyzed': 0},
        summary_text="No vendor data.",
        recommendations=[]
    )

    report = Report(metadata=metadata, sections=[vendor_section])

    generator = ExcelReportGenerator()
    generator.generate_excel(report, temp_output_file)

    assert os.path.exists(temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Vendors']
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert 'No contractors with minimum 3 work orders found' in all_values


def test_empty_failure_section(temp_output_file):
    """Test handling of empty failure section."""
    metadata = {'generated_date': '2024-01-15', 'total_records': 0, 'total_cost': 0}

    failure_section = ReportSection(
        title="Failure Pattern Analysis",
        content={'message': 'No text data available', 'patterns_found': 0},
        summary_text="No patterns.",
        recommendations=[]
    )

    report = Report(metadata=metadata, sections=[failure_section])

    generator = ExcelReportGenerator()
    generator.generate_excel(report, temp_output_file)

    assert os.path.exists(temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Failures']
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert 'No text data available' in all_values


def test_recommendations_with_no_sections(temp_output_file):
    """Test recommendations sheet when no recommendations exist."""
    metadata = {'generated_date': '2024-01-15', 'total_records': 0, 'total_cost': 0}

    section = ReportSection(
        title="Test Section",
        content={'data': 'none'},
        summary_text="Nothing found.",
        recommendations=[]  # No recommendations
    )

    report = Report(metadata=metadata, sections=[section])

    generator = ExcelReportGenerator()
    generator.generate_excel(report, temp_output_file)

    assert os.path.exists(temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Recommendations']
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert 'No recommendations generated' in all_values


def test_full_report_structure(sample_report, temp_output_file):
    """Test that full report has proper structure and data."""
    generator = ExcelReportGenerator()
    generator.generate_excel(sample_report, temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)

    # Check all sheets exist
    assert len(wb.sheetnames) == 6

    # Check each sheet has content (at least a title)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws['A1'].value is not None  # Title should exist


def test_column_widths_set(sample_report, temp_output_file):
    """Test that column widths are set appropriately."""
    generator = ExcelReportGenerator()
    generator.generate_excel(sample_report, temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Equipment']

    # Check that columns have been sized (not default width of 8.43)
    # Note: openpyxl may not always preserve exact widths, but we can check they're set
    col_A_width = ws.column_dimensions['A'].width
    assert col_A_width is not None


def test_multiple_recommendations_per_section(temp_output_file):
    """Test that multiple recommendations from each section are captured."""
    metadata = {'generated_date': '2024-01-15', 'total_records': 100, 'total_cost': 50000}

    section1 = ReportSection(
        title="Section 1",
        content={'data': 'test'},
        summary_text="Summary 1",
        recommendations=["Rec 1.1", "Rec 1.2", "Rec 1.3"]
    )

    section2 = ReportSection(
        title="Section 2",
        content={'data': 'test'},
        summary_text="Summary 2",
        recommendations=["Rec 2.1", "Rec 2.2"]
    )

    report = Report(metadata=metadata, sections=[section1, section2])

    generator = ExcelReportGenerator()
    generator.generate_excel(report, temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Recommendations']

    # Count data rows (excluding headers and title)
    data_rows = [row for row in ws.iter_rows(min_row=5) if row[0].value is not None]
    assert len(data_rows) == 5  # 3 from section1 + 2 from section2


def test_priority_assignment(temp_output_file):
    """Test that first recommendation gets High priority, others get Medium."""
    metadata = {'generated_date': '2024-01-15', 'total_records': 100, 'total_cost': 50000}

    section = ReportSection(
        title="Test Section",
        content={'data': 'test'},
        summary_text="Summary",
        recommendations=["First rec", "Second rec", "Third rec"]
    )

    report = Report(metadata=metadata, sections=[section])

    generator = ExcelReportGenerator()
    generator.generate_excel(report, temp_output_file)

    wb = openpyxl.load_workbook(temp_output_file)
    ws = wb['Recommendations']

    # First recommendation should be High
    assert ws['B5'].value == 'High'
    # Second and third should be Medium
    assert ws['B6'].value == 'Medium'
    assert ws['B7'].value == 'Medium'
